import pandas as pd
import requests
import os
from tqdm import tqdm  # Progress bar
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Function to get lat/lng from Plus Code using Google Maps API
def get_lat_lng_from_plus_code(plus_code, api_key):
    url = f"https://maps.googleapis.com/maps/api/geocode/json?address={plus_code}&key={api_key}"
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        if 'results' in data and len(data['results']) > 0:
            location = data['results'][0]['geometry']['location']
            return location['lat'], location['lng']
    return None, None

# Function to auto-adjust the column widths for an XLSX file
def adjust_column_width(output_file_path):
    wb = load_workbook(output_file_path)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid cells that might be None
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    wb.save(output_file_path)

# Ask for the API key (set manually in your case)
api_key = "YOUR_GOOGLE_MAPS_API_KEY"

# Step 1: Ask user for the CSV or XLSX file using a file dialog
root = Tk()
root.withdraw()  # Hide the main Tkinter window
input_file_path = askopenfilename(
    title="Select CSV or XLSX file",
    filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx")] 
)

if not input_file_path:
    print("No file selected. Exiting.")
    exit()

# Step 2: Read the input file into DataFrame based on its extension
file_extension = os.path.splitext(input_file_path)[1].lower()

if file_extension == '.csv':
    df = pd.read_csv(input_file_path)
elif file_extension == '.xlsx':
    df = pd.read_excel(input_file_path)
else:
    raise Exception("Unsupported file format. Please select a CSV or XLSX file.")

# Step 3: Find the Plus Code column (adjust as needed)
plus_code_column = None
for col in df.columns:
    if "Plus" in col or "plus" in col:  # assuming the Plus Code column contains the word "Plus"
        plus_code_column = col
        break

if plus_code_column is None:
    raise Exception("Could not find Plus Code column in the CSV file.")

# Step 4: Generate lat and lng for all Plus Codes with progress bar
latitudes = []
longitudes = []

# Initialize progress bar using tqdm
for index, row in tqdm(df.iterrows(), desc="Converting Plus Codes", unit="code", total=len(df)):
    plus_code = row[plus_code_column]
    lat, lng = get_lat_lng_from_plus_code(plus_code, api_key)
    latitudes.append(lat)
    longitudes.append(lng)

# Step 5: Add lat/lng columns to DataFrame
df['Latitude'] = latitudes
df['Longitude'] = longitudes

# Step 6: Create 'with-lat-lng' folder if it doesn't exist
output_dir = "with-lat-lng"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Step 7: Ask the user for the desired output format (1 for csv, 2 for xlsx)
while True:
    print("\nChoose output format:")
    print("1 - CSV")
    print("2 - XLSX")
    user_choice = input("Enter 1 or 2: ").strip()

    if user_choice == '1':
        output_format = 'csv'
        break
    elif user_choice == '2':
        output_format = 'xlsx'
        break
    else:
        print("Invalid choice. Please select 1 or 2.")

# Step 8: Save the file in the chosen format
output_file_name = os.path.splitext(os.path.basename(input_file_path))[0] + "_with_lat_lng"
output_file_path = os.path.join(output_dir, f"{output_file_name}.{output_format}")

if output_format == 'xlsx':
    # Save as .xlsx
    df.to_excel(output_file_path, index=False)

    # Adjust the column width for better readability
    adjust_column_width(output_file_path)
    print(f"File saved successfully as {output_file_path} (with auto-adjusted column widths).")

elif output_format == 'csv':
    # Save as .csv
    df.to_csv(output_file_path, index=False)
    print(f"File saved successfully as {output_file_path}.")
